"""
Export service: Generate PDF and Excel reports from analysis results.
"""
import io
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie as RLPie


def generate_excel(results_data: dict) -> bytes:
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, cols):
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    analysis = results_data.get("analysis", {})
    results = results_data.get("results", {})

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    style_header(ws_summary, ["Metric", "Value"])
    summary_rows = [
        ("Filename", analysis.get("filename", "")),
        ("Total Reviews", analysis.get("total_reviews", 0)),
        ("Best Model", results.get("best_model", "")),
        ("Model Accuracy", f"{results.get('best_accuracy', 0) * 100:.1f}%"),
        ("Positive Reviews", results.get("sentiment_distribution", {}).get("positive", 0)),
        ("Negative Reviews", results.get("sentiment_distribution", {}).get("negative", 0)),
        ("Neutral Reviews", results.get("sentiment_distribution", {}).get("neutral", 0)),
        ("Problems Detected", len(results.get("problems", {}).get("problems", []))),
        ("Recommendations", results.get("recommendations", {}).get("total_recommendations", 0)),
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for row_idx, (metric, value) in enumerate(summary_rows, 2):
        ws_summary.cell(row=row_idx, column=1, value=metric).border = border
        ws_summary.cell(row=row_idx, column=2, value=value).border = border
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 35

    # Sheet 2: Sentiment Distribution
    ws_sent = wb.create_sheet("Sentiment")
    dist = results.get("sentiment_distribution", {})
    style_header(ws_sent, ["Sentiment", "Count", "Percentage"])
    total = dist.get("total", 1)
    for row_idx, sentiment in enumerate(["positive", "negative", "neutral"], 2):
        count = dist.get(sentiment, 0)
        pct = f"{(count / total * 100):.1f}%"
        ws_sent.cell(row=row_idx, column=1, value=sentiment.capitalize()).border = border
        ws_sent.cell(row=row_idx, column=2, value=count).border = border
        ws_sent.cell(row=row_idx, column=3, value=pct).border = border

    # Pie chart
    pie = PieChart()
    pie.title = "Sentiment Distribution"
    pie.style = 10
    data_ref = Reference(ws_sent, min_col=2, min_row=1, max_row=4)
    cats_ref = Reference(ws_sent, min_col=1, min_row=2, max_row=4)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)
    ws_sent.add_chart(pie, "E2")

    # Sheet 3: Problems
    ws_prob = wb.create_sheet("Problems")
    style_header(ws_prob, ["Category", "Severity", "Frequency", "Percentage", "Examples"])
    problems = results.get("problems", {}).get("problems", [])
    for row_idx, p in enumerate(problems, 2):
        ws_prob.cell(row=row_idx, column=1, value=p.get("category", "")).border = border
        ws_prob.cell(row=row_idx, column=2, value=p.get("severity", "")).border = border
        ws_prob.cell(row=row_idx, column=3, value=p.get("frequency", 0)).border = border
        ws_prob.cell(row=row_idx, column=4, value=f"{p.get('percentage', 0)}%").border = border
        examples = "; ".join(p.get("examples", [])[:2])
        ws_prob.cell(row=row_idx, column=5, value=examples[:200]).border = border
    ws_prob.column_dimensions["E"].width = 60

    # Sheet 4: Recommendations
    ws_rec = wb.create_sheet("Recommendations")
    style_header(ws_rec, ["Title", "Priority", "Category", "Impact", "Suggestions"])
    recs = results.get("recommendations", {}).get("recommendations", [])
    for row_idx, r in enumerate(recs, 2):
        ws_rec.cell(row=row_idx, column=1, value=r.get("title", "")).border = border
        ws_rec.cell(row=row_idx, column=2, value=r.get("priority", "")).border = border
        ws_rec.cell(row=row_idx, column=3, value=r.get("problem_category", "")).border = border
        ws_rec.cell(row=row_idx, column=4, value=r.get("impact", "")[:200]).border = border
        suggestions = "; ".join(r.get("suggestions", [])[:3])
        ws_rec.cell(row=row_idx, column=5, value=suggestions[:300]).border = border
    ws_rec.column_dimensions["D"].width = 40
    ws_rec.column_dimensions["E"].width = 50

    # Sheet 5: Model Comparison
    ws_model = wb.create_sheet("Models")
    style_header(ws_model, ["Model", "Accuracy"])
    model_results = results.get("model_results", {})
    for row_idx, (name, info) in enumerate(model_results.items(), 2):
        ws_model.cell(row=row_idx, column=1, value=name.replace("_", " ").title()).border = border
        ws_model.cell(row=row_idx, column=2, value=f"{info.get('accuracy', 0) * 100:.1f}%").border = border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_pdf(results_data: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle("Title2", parent=styles["Title"], fontSize=20, textColor=colors.HexColor("#6366F1"))
    heading_style = ParagraphStyle("Heading", parent=styles["Heading2"], fontSize=14, textColor=colors.HexColor("#334155"))
    body_style = ParagraphStyle("Body2", parent=styles["Normal"], fontSize=10, leading=14)
    small_style = ParagraphStyle("Small", parent=styles["Normal"], fontSize=8, textColor=colors.grey)

    analysis = results_data.get("analysis", {})
    results = results_data.get("results", {})

    elements.append(Paragraph("AnZlyze - Analysis Report", title_style))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", small_style))
    elements.append(Spacer(1, 20))

    # Summary
    elements.append(Paragraph("Summary", heading_style))
    dist = results.get("sentiment_distribution", {})
    total = dist.get("total", 1)
    summary_data = [
        ["Metric", "Value"],
        ["Filename", analysis.get("filename", "")],
        ["Total Reviews", str(analysis.get("total_reviews", 0))],
        ["Best Model", results.get("best_model", "").replace("_", " ").title()],
        ["Accuracy", f"{results.get('best_accuracy', 0) * 100:.1f}%"],
        ["Positive", f"{dist.get('positive', 0)} ({dist.get('positive', 0) / total * 100:.1f}%)"],
        ["Negative", f"{dist.get('negative', 0)} ({dist.get('negative', 0) / total * 100:.1f}%)"],
        ["Neutral", f"{dist.get('neutral', 0)} ({dist.get('neutral', 0) / total * 100:.1f}%)"],
    ]
    t = Table(summary_data, colWidths=[2.5 * inch, 3.5 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6366F1")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 20))

    # Problems
    problems = results.get("problems", {}).get("problems", [])
    if problems:
        elements.append(Paragraph("Detected Problems", heading_style))
        prob_data = [["Category", "Severity", "Frequency", "Percentage"]]
        for p in problems:
            prob_data.append([
                p.get("category", ""),
                p.get("severity", "").upper(),
                str(p.get("frequency", 0)),
                f"{p.get('percentage', 0)}%",
            ])
        t2 = Table(prob_data, colWidths=[2 * inch, 1.2 * inch, 1.2 * inch, 1.2 * inch])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6366F1")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(t2)
        elements.append(Spacer(1, 20))

    # Recommendations
    recs = results.get("recommendations", {}).get("recommendations", [])
    if recs:
        elements.append(Paragraph("Recommendations", heading_style))
        for r in recs:
            priority = r.get("priority", "").upper()
            elements.append(Paragraph(f"[{priority}] {r.get('title', '')}", body_style))
            elements.append(Paragraph(f"<i>{r.get('impact', '')}</i>", small_style))
            for s in r.get("suggestions", [])[:3]:
                elements.append(Paragraph(f"  - {s}", body_style))
            elements.append(Spacer(1, 8))

    doc.build(elements)
    buf.seek(0)
    return buf.read()
